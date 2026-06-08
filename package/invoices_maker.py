from openpyxl import load_workbook
import os, sys
sys.path.append(os.getcwd())
from pprint import pprint

from package.config import BANK_ACCOUNT_ID
from package.api import create_invoice

from datetime import datetime

def negative_weights_detector():
        
        wb= load_workbook(os.path.join(os.getcwd(), 'Для накладных.xlsx'))
        ws = wb.active

        for row_number in range(4, ws.max_row-1):
            row_cells = ws[row_number]
            row_values = [cell.value for cell in row_cells] 

            if '-' in str(row_values[6]):
                raise ValueError('Столбец G "Масса нетто" содержит отрицательные значения!')


def invoices_maker():
    try:
        negative_weights_detector()
        
        wb= load_workbook(os.path.join(os.getcwd(), 'Для накладных.xlsx'))
        ws = wb.active
        warehouseItems = []
        if ws.cell(column=1, row=3).value != 'Накладных:':
            return (False, 'Файл "Для накладных.xlsx" содержит ошибки')
        else:
            invoice_qty = str(ws.cell(column=2, row=3).value)
        
        invoice_number = 0
        for row_number in range(5, ws.max_row):
            row_cells_values = [str(cell.value) for cell in ws[row_number]]

            if row_cells_values[3] == 'Закрытие партии/чека':
                invoice_number += 1
                print(f'{invoice_number} из {invoice_qty}. Создаём накладную для партии {product_batch}')
                # pprint(payload)
                create_invoice(payload)
                warehouseItems = []
            else:
                if row_cells_values[8]!='0':
                    unitName = "шт"
                    quantity = float(row_cells_values[8])
                else:
                    unitName = "кг"
                    quantity = float(row_cells_values[6])
                        
                warehouseItem = {
                    "productName": row_cells_values[5],
                    "unitName": unitName,
                    "quantity": quantity,
                    "price": float(row_cells_values[9]),
                    "ndsRate": None,
                    "discount": None
                    }

                warehouseItem = {
                    "productName": row_cells_values[5],
                    "unitName": "кг",
                    "quantity": float(row_cells_values[6]),
                    "price": float(row_cells_values[9]),
                    "ndsRate": None,
                    "discount": None
                    }
                                
                warehouseItems.append(warehouseItem)
                
                product_batch = row_cells_values[13]
                
                payload = {
                    "request": {},
                    "bankAccountId": BANK_ACCOUNT_ID,
                    "contractorId": row_cells_values[18],
                    "consigneeId": row_cells_values[18],      
                    "number": product_batch,
                    "warehouseItems": warehouseItems,
                    "date": datetime.strptime(row_cells_values[2], "%d.%m.%Y %H:%M:%S").strftime("%Y-%m-%d")
                }
                
        return (True, 'Накладные успешно созданы')       
                
                
    except Exception as e:
        return (False, repr(e))

if __name__ == '__main__':
    # invoices_maker()
    negative_weights_detector()