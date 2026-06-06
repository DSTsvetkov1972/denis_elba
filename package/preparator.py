import pandas as pd
import os,sys
import random

sys.path.append(os.getcwd())

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from package.sqlite import get_contractor_id, get_good_name, get_contractor_name_from_idFromName

FONT_COLOR_GREEN = Font(color='2E8B57')
FONT_COLOR_RED = Font(color='FF0000')
LIGHT_BLUE_FILL = PatternFill(start_color='DDEBF7', 
                               end_color='DDEBF7', 
                               fill_type='solid')


def check_source_folder():
    source_files = list(os.walk('Исходник'))[0][2]
    source_files_qty = len(source_files)

    if source_files_qty!=1:
        return (False, f'В папке Исходник должен быть только один файл. Сейчас там: {source_files_qty}')
    elif source_files[0][-4:] != 'xlsx':
        return (False, 'В папке Исходник должен находится файл с расширением .xlsx')
    else:
        return (True, source_files[0])
    
def get_light_hex():
    r = random.randint(200, 255)
    g = random.randint(160, 255)
    b = random.randint(100, 200)
    return f"{r:02x}{g:02x}{b:02x}"    

def download_file_preparator(source_file_name):
    
    #try: 
    if 1:
        invoices_qty = 0
        if os.path.exists(os.path.join(os.getcwd(), 'Для накладных.xlsx')):
            os.remove(os.path.join(os.getcwd(), 'Для накладных.xlsx'))
            
        wb = load_workbook(os.path.join(os.getcwd(), 'Исходник', source_file_name))
        ws = wb.active
        contractors_errors = 0
        goods_errors = 0
        repetitions_dict = {}

        for row_number in range(ws.max_row-1, 4, -1):
                 
            row_cells = ws[row_number]
            row_values = [cell.value for cell in row_cells]
            
            article = row_values[4]
            product_batch = row_values[13]
            operation = row_values[3]

            if row_number>5:
                preceding_idFromName = ws.cell(column=16, row=row_number-1).value
            else:     
                preceding_idFromName = None  
            get_contractor_res = get_contractor_name_from_idFromName(preceding_idFromName)    
            # contractor = row_values[18] #get_contractor_name_from_idFromName(idFromName)
            contractor_cell = ws.cell(column=19, row=row_number)

            if operation in ['Закрытие смены', 'Аварийное закрытие партии']:
                for column_number in range(1, ws.max_column+1):
                    ws.cell(row=row_number, column=column_number).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws.cell(row=row_number, column=column_number).font = Font(color="FFFFFF", bold=True)

            if operation in ['Закрытие партии/чека']:
                for column_number in range(1, ws.max_column+1):
                    ws.cell(row=row_number, column=column_number).fill = PatternFill(start_color="707070", end_color="707070", fill_type="solid")
                    ws.cell(row=row_number, column=column_number).font = Font(color="FFFFFF", bold=True)   

                if not get_contractor_res[0]:
                    contractor_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    
                contractor = get_contractor_res[1]     
                contractor_cell.value = contractor 
                
            else:
                contractor = None   

            if operation in ['Закрытие партии/чека', 'Закрытие смены', 'Аварийное закрытие партии']:
                invoices_qty += 1
                
                contractor_id_res = get_contractor_id(contractor, operation)
                if contractor_id_res[0]:
                    contractor_cell_color = FONT_COLOR_GREEN
                else:
                    contractors_errors += 1
                    contractor_cell_color = FONT_COLOR_RED
                contractor_cell_value = contractor_id_res[1]
                
            elif article:
                # собираем в словарь информацию о повторении артикула внутри партии товара

                repetitions_dict_key = (product_batch, article)
                if repetitions_dict_key in repetitions_dict:
                    repetitions_dict[repetitions_dict_key].append(row_number)
                else:
                    repetitions_dict[repetitions_dict_key] = [row_number]

                contractor_cell = ws.cell(row=row_number, column=19)
                contractor_cell.value = contractor_cell_value
                contractor_cell.font = contractor_cell_color
                
                productMainName_cell = ws.cell(row=row_number, column=6)
                get_good_name_res = get_good_name(row_values[4])
                
                if get_good_name_res[0]:
                    ws.cell(row=row_number, column=5).font = FONT_COLOR_GREEN
                    productMainName_cell.font = FONT_COLOR_GREEN
                else:
                    goods_errors += 1
                    ws.cell(row=row_number, column=5).font = FONT_COLOR_RED
                    productMainName_cell.font = FONT_COLOR_RED
                    
                productMainName_cell.value = get_good_name_res[1]

        # подсвечиваем строки с повторяющимися артикулами внутри партии
        #print(repetitions_dict)
        for repetitions_dict_key, row_numbers in repetitions_dict.items():
            if len(row_numbers)>1:
                reapiting_cell_color =  get_light_hex()
                reapiting_cell_fill = PatternFill(start_color=reapiting_cell_color, end_color=reapiting_cell_color, fill_type="solid")
                for row_number in row_numbers:
                    for column_number in range(1, ws.max_column+1):
                        reapiting_cell = ws.cell(column=column_number, row=row_number)
                        reapiting_cell.fill = reapiting_cell_fill

        # подсвечиваем красным ячейки с отрицательным весом
        for row_number in range(ws.max_row-1, 4, -1):
            row_cells = ws[row_number]
            row_values = [str(cell.value) for cell in row_cells]
            if row_values[6] and '-' in row_values[6]:
                ws.cell(column=7, row=row_number).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
        
        err_message_cell = ws.cell(row=3, column=1)
        invoices_qty_cell = ws.cell(row=3, column=2)
        
        if contractors_errors or goods_errors:
            err_message = f"Не удалось определить конрагентов: {contractors_errors}; товаров: {goods_errors}"
            err_message_cell.font = FONT_COLOR_RED
            err_message_cell.value = err_message
            wb.save('Для накладных.xlsx')
            return (False, err_message)
        else:
            err_message_cell.font = FONT_COLOR_GREEN
            err_message_cell.value = 'Накладных:'
            
            invoices_qty_cell.font = FONT_COLOR_GREEN
            invoices_qty_cell.value = invoices_qty
            
            wb.save('Для накладных.xlsx')
            return (True, 'Файл "Для накладных.xlsx" подготовлен. Ассортимент и контрагенты сопоставлены успешно.')
    
    #except Exception as e:
    #    return(False, f"download_file_preparator {repr(e)}")

    
if __name__ == '__main__':
    print(download_file_preparator('05.06.26 тест.xlsx'))
