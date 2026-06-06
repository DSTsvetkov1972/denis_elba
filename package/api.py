
import os, sys, shutil
sys.path.append(os.getcwd())

import requests
import pandas as pd
from colorama import init, Fore
from package.config import API_KEY, COMPANY_ID, BANK_ACCOUNT_ID
from progress.bar import FillingSquaresBar
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

session = requests.Session()
session.keep_alive = False  # отключаем постоянные соединени
# ConnectionError(ProtocolError('Connection aborted.', ConnectionResetError(10054, 'Удаленный хост принудительно разорвал существующее подключение', None, 10054, None)))

def get_article_from_name(name):
    
    name_parts = str(name).split(' ')

    if len(name_parts) > 1:
        try:
            return int(name_parts[-1])
        except ValueError:
            return None 
    else:
        return None
             
    
    

def excel_formater(excel_file_path, column_widths):
    wb = load_workbook(excel_file_path)
    ws = wb.active

    ws.freeze_panes = 'A2'

    for n, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(n)].width = column_width
        ws.cell(row=1, column=n).font = Font(bold=True)

    wb.save(excel_file_path)





def get_all_goods():
    try:
        if os.path.exists(os.path.join(os.getcwd(), 'Номенклатура.xlsx')):
            os.remove(os.path.join(os.getcwd(), 'Номенклатура.xlsx'))

        url = f"https://elba-api.kontur.ru/v1/organizations/{COMPANY_ID}/products/search?"
        headers = {
            "X-Kontur-ApiKey": API_KEY,
            "Content-Type": "application/json",
            'Connection': 'close'
            }
        dfs = []
        
        bar = FillingSquaresBar(
            ' Выгружаем ассортимент:',
            max=10,
            suffix = '%(index)d/%(max)d',
            fill='█', empty_fill='░',
            width = 50)  
        

        bar.start()
        for fig in range(0, 10):
            payload = {
                "filter": {
                    "article": str(fig),                                              
                }
            }

            response = session.post(url, headers=headers, json=payload)

            df = pd.DataFrame(response.json()['products'])
            dfs.append(df)
            
            bar.next()
            
        bar.finish()    
            
        goods_df = pd.concat(dfs, ignore_index=True)
        goods_df.drop_duplicates(inplace=True)
        goods_df = goods_df.sort_values('productMainName')


        goods_df.to_excel(os.path.join(os.getcwd(), 'Номенклатура.xlsx'), index=None)

        excel_formater(
            os.path.join(os.getcwd(), 'Номенклатура.xlsx'),
            [38, 32, 8]
        )
        
        return (True, goods_df)
    
    except Exception as e:
        return (False, repr(e))
    

def get_all_contractors():
    
    try:
        if os.path.exists(os.path.join(os.getcwd(), 'Контрагенты.xlsx')):
            os.remove(os.path.join(os.getcwd(), 'Контрагенты.xlsx'))

        url = f"https://elba-api.kontur.ru/v1/organizations/{COMPANY_ID}/contractors/search?"

        headers = {
            "X-Kontur-ApiKey": API_KEY,
            "Content-Type": "application/json",
            'Connection': 'close'
            }
        dfs = []
        
        bar = FillingSquaresBar(
            'Выгружаем контрагентов:',
            max=33,
            suffix = '%(index)d/%(max)d',
            fill='█', empty_fill='░',
            width = 50)  
        
        bar.start()
        
        for letter in 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя':

            payload = {
                "filter": {
                    "name": letter,                                              
                }
            }

            response = session.post(url, headers=headers, json=payload)
            
            if response.status_code != 200:
                print(response.status_code)
            
            df = pd.DataFrame(response.json()['contractors'])
            dfs.append(df)
            bar.next()
            
        bar.finish()    
            
        contractors_df = pd.concat(dfs, ignore_index=True)
        contractors_df.drop_duplicates(inplace=True)
        contractors_df = contractors_df.sort_values('name')
        contractors_df['idFromName'] = contractors_df['shortName'].apply(get_article_from_name)
        contractors_df = contractors_df[['id', 'name', 'shortName', 'idFromName', 'inn', 'kpp', 'address']]

        contractors_df.to_excel(os.path.join(os.getcwd(), 'Контрагенты.xlsx'), index=None)

        excel_formater(
            os.path.join(os.getcwd(), 'Контрагенты.xlsx'),
            [38, 42, 24, 12, 12, 12, 42]
        )
        
        return (True, contractors_df)
    
    except Exception as e:
        return (False, repr(e))


def create_invoice(payload):
    
    # url = f"https://elba-api.kontur.ru/v1/organizations/{COMPANY_ID}/products/search?"
    url = f"https://elba-api.kontur.ru/v1/organizations/{COMPANY_ID}/delivery-notes?"
    headers = {
        "X-Kontur-ApiKey": API_KEY,
        "Content-Type": "application/json",
        'Connection': 'close'
        }
    
    payload = payload
    
    response = requests.post(url, headers=headers, json=payload)
    
    if response.status_code != 201:
        raise Exception(response.json())
        

    
    
if __name__ == '__main__':
    payload = {
        "request": {},
        "bankAccountId": BANK_ACCOUNT_ID,
        "contractorId": "443a7b62-48ab-45a2-ba25-53adbf1edceb",
        "consigneeId": "443a7b62-48ab-45a2-ba25-53adbf1edceb",
        "date": "2026-05-31",
        'warehouseItems': [
            {'discount': None,
             'ndsRate': None,
             'price': 365,
             'productName': 'Лук зеленый',
             'quantity': 3.3,
             'unitName': 'кг'}]
    }
 
    create_invoice(payload)    
    

if __name__ == '__main__':
    print(get_article_from_name('Dfcz 11'))