from datetime import datetime
import os, shutil
from openpyxl import load_workbook

def confirm():
    try:
    # if 1:

        wb= load_workbook(os.path.join(os.getcwd(), 'Для накладных.xlsx'))
        ws = wb.active
        if ws.cell(column=1, row=3).value != 'Накладных:':
            return (False, 'Файл "Для накладных.xlsx" содержит ошибки')
        
        archive_datetime = str(datetime.now())[:19].replace(':', '-')
     
        shutil.move(
            os.path.join(os.getcwd(), 'Для накладных.xlsx'),
            os.path.join(os.getcwd(), 'Архив', f"{archive_datetime}_Для накладных.xlsx")
        )

        source_file_name = list(os.walk(os.path.join(os.getcwd(), 'Исходник')))[0][2][0]
        source_file_path = os.path.join(os.getcwd(), 'Исходник', source_file_name)
        shutil.move(
            source_file_path,
            os.path.join(os.getcwd(), 'Архив', f"{archive_datetime}_{source_file_name}"))
        
        return (True, f'Файлы "Для накладной.xlsx" и {archive_file_name} перемещены в архив')
    except Exception as e:
        return (False, repr(e))
    
if __name__ == '__main__':
    print(confirm())